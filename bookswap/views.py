from django.shortcuts import render, HttpResponseRedirect
from django.contrib import messages
from django.db import connection
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
import requests
from .forms import user_details, login_details, book_search, email_update, mobile_update, pass_input
from django.views.decorators.csrf import csrf_protect
from django.core.mail import send_mail
import MySQLdb
import urllib2
import urllib
import urlparse
import xmltodict
import json
import random

def sendSMS(apikey, numbers, sender, message):
    data =  urllib.urlencode({'apikey': apikey, 'numbers': numbers,
        'message' : message, 'sender': sender})
    data = data.encode('utf-8')
    request = urllib2.Request("https://api.textlocal.in/send/?")
    f = urllib2.urlopen(request, data)
    fr = f.read()
    return(fr)

def fix(request):

	resp =  sendSMS('8nR3v2N0Mn8-kLkdIOfwurKVo16acNGOHC3v9stoF0', '7052378759', 'TXTLCL', 'This is your message')
	print (resp)

	# with connection.cursor() as cursor:
	# 	cursor.execute("UPDATE collection set available = 1")

	# return render(request, 'success.html')

def lradd(request):

	b1 = bookdatabystring("not a penny more")
	#b2 = bookdatabystring("THE TITANS CURSE")
	s1 = "LR 1020"
	#s2 = "LR 1022"

	id1 = b1['GoodreadsResponse']['book']['id']
	n1 = b1['GoodreadsResponse']['book']['title']
	a1 = b1['GoodreadsResponse']['book']['authors']['author']
	try:
		a1 = a1['name']
	except:
		a1 = a1[0]['name']
	p1 = b1['GoodreadsResponse']['book']['num_pages']
	r1 = b1['GoodreadsResponse']['book']['average_rating']
	i1 = b1['GoodreadsResponse']['book']['image_url']
	n1 = n1.replace("'",r"\'")
	a1 = a1.replace("'",r"\'")

	if p1 == None:
		p1 = 0
	if r1 == None:
		r1 = 0

	with connection.cursor() as cursor:
		cursor.execute("INSERT INTO letsbooks VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s')" % (0,id1,n1,a1,p1,r1,i1,s1,0))
		#cursor.execute("INSERT INTO letsbooks VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s')" % (0,id2,n2,a2,p2,r2,i2,s2,0))

def bookdatabystring(title = ""):

	title = title.replace(" ", "+")
	string = "https://www.goodreads.com/book/title.xml?key=bDGHlS4YtKU5RK78QC33Q&title=" + title

	try:
		file = urllib2.urlopen(string)
		data = file.read()
		file.close()
		data = xmltodict.parse(data)
	except:
		data = None

	return data

def letsreadprepare(request):

	from openpyxl import load_workbook

	wb = load_workbook('booklist.xlsx')
	sheet = wb.get_sheet_by_name('Sheet1')

	for i in range(5,89):
		serial = sheet.cell(row=i, column=1).value 
		title = sheet.cell(row=i, column=2).value

		data = bookdatabystring(title)

		if data != None:
			ID = data['GoodreadsResponse']['book']['id']
			name = data['GoodreadsResponse']['book']['title']
			author = data['GoodreadsResponse']['book']['authors']['author']
			try:
				author = author['name']
			except:
				author = author[0]['name']
			pages = data['GoodreadsResponse']['book']['num_pages']
			rating = data['GoodreadsResponse']['book']['average_rating']
			img = data['GoodreadsResponse']['book']['image_url']
			name = name.replace("'",r"\'")
			author = author.replace("'",r"\'")

			if pages == None:
				pages = 0
			if rating == None:
				rating = 0

			with connection.cursor() as cursor:
				cursor.execute("INSERT INTO letsbooks VALUES ('%s','%s','%s','%s','%s','%s','%s','%s')" % (0,ID,name,author,pages,rating,img,serial))

def notify(uid):
	
	with connection.cursor() as cursor:
		cursor.execute("SELECT * from requests where toid = '%s' AND status = 0" % (uid))
		check = cursor.fetchall()

	if len(check) == 0:
		return False
	else:
		return True

def emailchange(uid, email):

	with connection.cursor() as cursor:
		cursor.execute("UPDATE user_detail set email = '%s' where userid = '%s'" % (email, uid))

def mass_mail(subject, message):
	
	with connection.cursor() as cursor:
		cursor.execute("SELECT email from user_detail")
		mail = cursor.fetchall()

	for i in mail:
		send_mail(subject, message, 'no-reply@bookfish.in', [i[0]], fail_silently = False,)

def delete_acc(username, uid):

	with connection.cursor() as cursor:
		cursor.execute("DELETE from requests where toid='%s' OR fromid='%s'" % (uid, uid))
		cursor.execute("DELETE from collection where userid='%s'" % (uid))
		cursor.execute("DELETE from wishlist where userid='%s'" % (uid))
		cursor.execute("DELETE from user_detail where userid='%s'" % (uid))
		cursor.execute("DELETE from auth_user where username='%s'" % (username))

def booknamebyid(id = 0):
	
	url = "https://www.goodreads.com/book/show/"+str(id)+".xml?key=bDGHlS4YtKU5RK78QC33Q"

	global data

	file = urllib2.urlopen(url)
	data = file.read()
	file.close()

	data = xmltodict.parse(data)
	result = data['GoodreadsResponse']['book']['title']

	return result

def booknamebystring(title = ""):

	title = title.replace(" ", "+")
	string = "https://www.goodreads.com/book/title.xml?key=bDGHlS4YtKU5RK78QC33Q&title=" + title

	global data

	try:
		file = urllib2.urlopen(string)
		data = file.read()
		file.close()
		data = xmltodict.parse(data)

		result = data['GoodreadsResponse']['book']['title']
	except:
		result = "Sorry, Book not found!"

	return result

def social_reg(request):

	un = request.user.get_username()

	if request.method == 'POST':
		name = request.POST.get('Name')
		contact = request.POST.get('Contact')
		address = request.POST.get('Address')
		email = request.POST.get('email')
		username = request.user.get_username()
		count = 0;

		with connection.cursor() as cursor:
			cursor.execute("""SELECT * from user_detail where email = '%s'"""%(email))
			check = cursor.fetchall()

		if len(check) != 0:
			u = User.objects.get(username = un)
			u.delete()
			return HttpResponseRedirect('/5/')

	    #Server Connection to MySQL:
		with connection.cursor() as cursor:
			cursor.execute("""INSERT INTO user_detail VALUES ('%s','%s','%s','%s','%s','%s','%s')"""%(name,0,contact,address,email,username,count))

		return HttpResponseRedirect('/login/')

def fb(request):

	un = request.user.get_username()

	with connection.cursor() as cursor:
		cursor.execute("""SELECT * from user_detail where username = '%s'"""%(un))
		exist = cursor.fetchall();

	if len(exist) == 0:
		form = user_details()

		user = User.objects.get(username=un)
		social = user.social_auth.get(provider='facebook')
		token = social.extra_data['access_token']

		string = "https://graph.facebook.com/me?fields=first_name,name,email&access_token=" + token
		file = urllib2.urlopen(string)
		data = file.read()
		file.close()

		data = json.loads(data)

		try:
			em = data['email']
			with connection.cursor() as cursor:
				cursor.execute("""SELECT * from user_detail where email = '%s'"""%(em))
				check = cursor.fetchall()
		except:
			check = ""

		if len(check) != 0:
			u = User.objects.get(username = un)
			u.delete()
			return HttpResponseRedirect('/5/')

		return render(request, 'fb.html', context = {"data":data, "form":form})
	else:
		return HttpResponseRedirect('/login/')

def acc_verify(request, uid = 0):

	if request.method == 'POST':

		form = pass_input(request.POST)

		if form.is_valid():
			password = form.cleaned_data['Password']			
			uid = form.cleaned_data['uid']

			try:
				with connection.cursor() as cursor:
					cursor.execute("SELECT username, email from user_detail where userid = '%s'" % (uid))
					temp = cursor.fetchall()
			except:
				return HttpResponseRedirect('/')

			username = temp[0][0]
			email = temp[0][1]

			user = User.objects.create_user(username, email, password)

			return HttpResponseRedirect('/4/')

	else:
		if uid == 0:
			return HttpResponseRedirect('/')
		else:
			uid = int(uid)
			uid = uid/243

			try:
				with connection.cursor() as cursor:
					cursor.execute("SELECT * from user_detail where userid = '%s'" % (uid))
					temp = cursor.fetchall()
			except:
				return HttpResponseRedirect('/')

			if not temp:
				return HttpResponseRedirect('/')				


		form = pass_input()
		return render(request, "verify.html", {'form':form, 'uid':uid})

def req_accept(request):

	reqid = request.POST.get('reqid')
	with connection.cursor() as cursor:
		cursor.execute("UPDATE requests set status=1 where reqid = '%s'" % reqid)
		cursor.execute("""SELECT fromid, bookreq, toid FROM requests where reqid = '%s' """ % (reqid))
		temp = cursor.fetchall()
		uid = temp[0][0]
		reqbook = temp[0][1]

		cursor.execute("""SELECT name, email, contact FROM user_detail where userid = '%s' """ % (uid))
		details = cursor.fetchall()
		name = details[0][0]
		email = details[0][1]
		cont = details[0][2]

		name = name.split(' ')
		name = name[0]

		cursor.execute("""SELECT name FROM collection where bookid = '%s' """ % (reqbook))
		bk_name = cursor.fetchall()
		bk_name = bk_name[0][0]

		if(temp[0][2] == 0):
			message = message = "Hello, " + name + "!"+ "\n \n" + "Your request for book " + "\"" + bk_name + "\"" + " has been accepted" + ". Login soon and contact Let's Read @ IIT-BHU. The contact is available in your requests tab.\n" + "\n" + "Cheers!\n" + "Team BookFish"
		else:
			message = "Hello, " + name + "!"+ "\n \n" + "Your request for book " + "\"" + bk_name + "\"" + " has been accepted" + ". Login soon and contact the owner.\n" +  "\n" + "Cheers!\n" + "Team BookFish" + "\n \n"

	send_mail('Book Request accepted', message, 'no-reply@bookfish.in', [email], fail_silently=True,)

	try:
		sendSMS('8nR3v2N0Mn8-kLkdIOfwurKVo16acNGOHC3v9stoF0', cont, 'TXTLCL', message)
	except:
		print "sms not sent"

	return HttpResponseRedirect('/requests/')

def req_reject(request):

	reqid = request.POST.get('reqid')
	with connection.cursor() as cursor:
		cursor.execute("UPDATE requests set status=-1 where reqid = '%s'" % reqid)

	return HttpResponseRedirect('/requests/')

def book_exists(table = "", bid = 0, uid = 0):

	if table == "collection":
		with connection.cursor() as cursor:
			cursor.execute("SELECT * from collection where bookid='%s' && userid='%s'" % bid,uid)
			check = cursor.fetchall()
	elif table == "wishlist":
		with connection.cursor() as cursor:
			cursor.execute("SELECT * from wishlist where bookid='%s' && userid='%s'" % bid,uid)
			check = cursor.fetchall()

	if len(check) == 0:
		return true
	else:
		return false

def idfrom_username(username = ""):

	with connection.cursor() as cursor:
		cursor.execute("SELECT userid from user_detail where username = '%s'" % username)
		uid = cursor.fetchall()
		uid = uid[0][0]
	return uid

def book_infcoll(username = ""):

	uid = idfrom_username(username)

	with connection.cursor() as cursor:
		cursor.execute("SELECT userid, bookid, name, author, img, available, entryid  from collection where userid = '%s'" % uid)
		book_info = cursor.fetchall()

	return book_info

def username_exist(username):
	
	with connection.cursor() as cursor:
		cursor.execute("SELECT * FROM user_detail where username='%s'" % username)
		check = cursor.fetchall()

	if len(check) != 0:
		return true
	else:
		return false

def changeAvailability(request):
	print "run"
	if(request.user.is_authenticated):
		if request.method == "POST":
			change = request.POST.get('num')
			eid = request.POST.get('entry')
			with connection.cursor() as cursor:
				cursor.execute("UPDATE collection set available = '%s' where entryid = '%s'" % (change, eid))
		else:
			HttpResponseRedirect('/')
	else:
		HttpResponseRedirect('/')

def available(request, sort="name"):

	if request.user.is_authenticated():
		username = request.user.get_username()
		uid = idfrom_username(username) 

		try:
			with connection.cursor() as cursor:
				if sort == "name" or sort == "author":
					cursor.execute("SELECT distinct bookid,name,author,pages,rating,img from collection where userid != '%s' AND available = 1 ORDER BY %s" % (uid, sort))
				else:
					cursor.execute("SELECT distinct bookid,name,author,pages,rating,img from collection where userid != '%s' AND available = 1 ORDER BY %s desc" % (uid, sort))

				books = cursor.fetchall()
		except:
			return HttpResponseRedirect('/')	

		notification = notify(idfrom_username(username))

		context = {'books':books, 'noti':notification}

		return render(request,"available.html",context)
	else:
		return HttpResponseRedirect('/')

def avail_owners(request, bid = 0):

	if request.user.is_authenticated():
		username = request.user.get_username()
		names = []

		if request.method == "POST":
			bookid = request.POST.get('own_book')
			print bookid
			uid = idfrom_username(username) 
			with connection.cursor() as cursor:
				cursor.execute("SELECT userid,book_cond from collection where bookid = '%s' AND userid != '%s' ORDER BY book_cond" % (bookid, uid))
				bookdata = cursor.fetchall()

				cursor.execute("SELECT name,bookid,userid from collection where userid = '%s'" % uid)
				collect = cursor.fetchall()

				for i in bookdata:
					cursor.execute("SELECT name from user_detail where userid = '%s'" % i[0])
					name = cursor.fetchall()
					names.append(name[0][0])
			
			x = len(bookdata)
			bookdata = list(bookdata)
			print names
			for j in xrange(0,x):
				bookdata[j] = list(bookdata[j])
			for j in xrange(0,x):
				bookdata[j].append(names[j])	

			notification = notify(idfrom_username(username))

			context = { 'owners':bookdata, 'collect':collect, 'reqbook':bookid, 'noti':notification }

			return render(request, "owners.html", context)

		else:
			return HttpResponseRedirect('/available/')

	else:
		return HttpResponseRedirect('/')

def account(request, change = ""):

	if request.user.is_authenticated():
		username = request.user.get_username()
		uid = idfrom_username(username)

		with connection.cursor() as cursor:
			cursor.execute("SELECT * FROM user_detail where userid = '%s'" % (uid))
			acc_det = cursor.fetchall()

		if request.method == 'POST':
			attr = request.POST.get('value')

			with connection.cursor() as cursor:
				cursor.execute("UPDATE user_detail set %s='%s' where userid = '%s'" % (change, attr, uid))

		notification = notify(idfrom_username(username))

		context = { 'det':acc_det, 'noti':notification }

		return render(request, "account.html", context)

	else:
		return HttpResponseRedirect('/')

# def mobile(request):

# 	if request.user.is_authenticated():
# 		username = request.user.get_username()
# 		uid = idfrom_username(username)

# 		if request.method == 'POST':
# 			form = mobile_update(request.POST)

# 			if form.is_valid():
# 				mobile = form.cleaned_data['mobile']

# 				print "mobile running"

# 				with connection.cursor() as cursor:
# 					cursor.execute("UPDATE user_detail set contact = '%s' where userid = '%s'" % (mobile, uid))

# 		else:
# 			email_form = email_update()
# 			mobile_form = mobile_update()
# 			return render(request, "account.html", {'emailform':email_form, 'mobileform':mobile_form})

# 	return HttpResponseRedirect('/')

# def mail(request):

# 	message = "Hello" + "!" + "\n" + "\n" + "You have pending book request(s). Please review them by logging in and get new books to read. Happy reading!" + "\n"+ "www.bookfish.in" +"\n" +"\n" + "Cheers!" + "\n" + "Team BookFish"
# 	subject = "Pending Requests"

# 	with connection.cursor() as cursor:
# 		cursor.execute("SELECT DISTINCT toid from requests where status=0")
# 		req = cursor.fetchall()

# 	with connection.cursor() as cursor:
# 		for i in req:
# 			cursor.execute("SELECT email from user_detail where userid= '%s'" % (i[0]))
# 			email = cursor.fetchall()
# 			email = email[0][0]

# 			send_mail(subject, message, 'no-reply@bookfish.in', [email], fail_silently = False,)

# 	return HttpResponseRedirect('/')		


def home(request, flag = 0):

	with connection.cursor() as cursor:
	 	cursor.execute("SELECT img, entryid FROM letsbooks")
	 	images = cursor.fetchall()

	empty = "nophoto"

	images = [x for x in images if not empty in x[0]]

	if len(images) > 60:
		images = random.sample(images, 60)

	# emails = []

	# for i in req:
	# 	if req[5] == 0:


	# subject = "Account Verification"

	# message1 = "Hello!"+ "\n \n" + "It looks like you faced some issues while email verification earlier. We are really sorry that it happened and will try our best to avoid such delays. Please click the link below to verify your account and explore the collection of books around you:" + "\n" + "www.bookfish.in/acc_verify/"
	# message2 = "\n \n" + "Cheers!\n" + "Team BookFish" + "\n \n" + "Click below to unsubscribe from these notifications (But, do you really want to?):"

	# message = message1 + str(29403) + message2
	# email = "amol.kankane@gmail.com"
	# send_mail(subject, message, 'no-reply@bookfish.in', [email], fail_silently = False,)

	if request.user.is_authenticated():
		logout(request)

	if request.method == "POST":
		form = user_details(request.POST)
		if form.is_valid():
			name = form.cleaned_data['Name']
			contact = form.cleaned_data['Mobile']
			address = form.cleaned_data['Address']
			email = form.cleaned_data['Email']
			username = form.cleaned_data['Username']
			password = form.cleaned_data['Password']
			confirmpassword = form.cleaned_data['ConfirmPassword']

			with connection.cursor() as cursor:
			 	cursor.execute("SELECT * FROM user_detail where username='%s'" % username)
			 	check = cursor.fetchall()

			flag = 2
			count = 0

		    #Server Connection to MySQL:
			with connection.cursor() as cursor:
			#conn = MySQLdb.connect(host= "localhost",user="root",passwd="root",db="startup")
			#x = conn.cursor()
				cursor.execute("""INSERT INTO user_detail VALUES ('%s','%s','%s','%s','%s','%s','%s')"""%(name,0,contact,address,email,username,count))
				cursor.execute("SELECT userid from user_detail where username = '%s'" % (username))
				uid = cursor.fetchall()
				uid = uid[0][0] * 243
				uid = str(uid)

			name = name.split(' ')
			name = name[0]

			link = "bookfish.au-syd.mybluemix.net/acc_verify/" + uid + "/"
			print link

			message = "Hello, " + name + "!"+ "\n \n" + "Click on this Link to complete your registration: " + "\n" + link + "\n \n" + "Cheers!\n" + "Team BookFish"
			
			send_mail('Account Activation', message, 'no-reply@bookfish.in', [email], fail_silently=False,)

			# except:
			# 	conn.rollback()
			form1 = user_details()
			form2 = login_details()
			
			context = { "sign_form":form, "log_form":form2, "flag":flag, "back":images}

		else:
			flag = 3
			form1 = user_details()
			form2 = login_details()
			context = { "sign_form":form, "log_form":form2, "flag":flag, "back":images}

	else:
		form1 = user_details()
		form2 = login_details()
		context = { "sign_form":form1, "log_form":form2, "flag":flag, "back":images}

	return render(request,"home.html",context)

@csrf_protect
def log_in(request, add = 0):

	flag = 0

	# Book add collection

	if request.method == "POST":
		form1 = login_details(request.POST)
		form2 = book_search(request.POST)

		if add == '1':
			if request.method == "POST":
				ID = data['GoodreadsResponse']['book']['id']
				name = data['GoodreadsResponse']['book']['title']
				author = data['GoodreadsResponse']['book']['authors']['author']
				try:
					author = author['name']
				except:
					author = author[0]['name']
				pages = data['GoodreadsResponse']['book']['num_pages']
				rating = data['GoodreadsResponse']['book']['average_rating']
				img = data['GoodreadsResponse']['book']['image_url']
				name = name.replace("'",r"\'")
				author = author.replace("'",r"\'")

				username = request.user.get_username()
				condition = request.POST.get('cond')

				#Server Connection to MySQL:
				uid = idfrom_username(username)

				with connection.cursor() as cursor:
					cursor.execute("SELECT * from collection where userid='%s' AND bookid='%s'" % (uid,ID))
					check = cursor.fetchall()
					if len(check) != 0:
						flag = 2
					else:
						try:
							cursor.execute("INSERT INTO collection VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s', '%s')" % (0,uid,ID,name,condition,author,pages,rating,img, 1))
						except:
							HttpResponseRedirect('/')
				add = 0

		if form1.is_valid():
			username = form1.cleaned_data['Username']
			password = form1.cleaned_data['Password']

			user = authenticate(username=username, password=password)

			if user == None:
				return HttpResponseRedirect('/1/')

			with connection.cursor() as cursor:
				cursor.execute("SELECT login from user_detail where username = '%s'" % (username))
				count = cursor.fetchall()
				count = count[0][0]
				count = count + 1

				cursor.execute("UPDATE user_detail set login = '%s' where username = '%s'" % (count, username))


			login(request, user)
			form = book_search()
			image = ""
			result = ""
			author = ""

		elif form2.is_valid():
			title = form2.cleaned_data['Title']
			title = booknamebystring(title)

			if title == "Sorry, Book not found!":
				flag = 1
				author = ""
				image = ""
			else:
				image = data['GoodreadsResponse']['book']['image_url']
				author = data['GoodreadsResponse']['book']['authors']['author']
				
				try:
					author = author['name']
				except:
					author = author[0]['name']

			form = book_search()
			result = title
		else:
			image = ""
			result = ""
			author = ""
			form = book_search()

		username = request.user.get_username()
		book_info = book_infcoll(username)
		notification = notify(idfrom_username(username))

		context = { 'form':form, 'image':image, 'title':result, 'author':author, 'books':book_info, 'flag':flag, 'noti':notification }

		return render(request, "login.html", context)
	else:

		if request.user.is_authenticated():
			username = request.user.get_username()

			with connection.cursor() as cursor:
				cursor.execute("SELECT * from user_detail where username = '%s'" % (username))
				check = cursor.fetchall()

			if len(check) == 0:
				return HttpResponseRedirect('/')

			form = book_search()
			image = ""
			result = ""
			author = ""
			book_info = book_infcoll(username)
			notification = notify(idfrom_username(username))

			context = { 'form':form,'image':image, 'title':result, 'author':author, 'books':book_info, 'flag':flag, 'noti':notification }

			return render(request, "login.html", context)
		else:
			return HttpResponseRedirect('/')

def wishlist(request, add=0):

	flag = 0

	# Adding book
	if add == '1':
		print("running")
		ID = data['GoodreadsResponse']['book']['id']
		img = data['GoodreadsResponse']['book']['image_url']

		author = data['GoodreadsResponse']['book']['authors']['author']		
		try:
			author = author['name']
		except:
			author = author[0]['name']

		name = data['GoodreadsResponse']['book']['title']
		name = name.replace("'",r"\'")
		username = request.user.get_username()

		print(name)

		#Server Connection to MySQL:

		uid = idfrom_username(username)
		with connection.cursor() as cursor:

			with connection.cursor() as cursor:
				cursor.execute("SELECT * from wishlist where userid='%s' AND bookid='%s'" % (uid,ID))
				check = cursor.fetchall()
				if len(check) != 0:
					flag = 2
				else:
					cursor.execute("INSERT INTO wishlist VALUES ('%s','%s','%s','%s','%s','%s')" % (0,uid,ID,name,author,img))

		add = 0

	if request.method == "POST":
		form = book_search(request.POST)

		if form.is_valid():
			username = request.user.get_username()

			title = form.cleaned_data['Title']
			result = booknamebystring(title)

			# Book search fails
			if result == "Sorry, Book not found!":
				img = ""
				author = ""
				flag = 1
			else:
				img = data['GoodreadsResponse']['book']['image_url']
				author = data['GoodreadsResponse']['book']['authors']['author']	
				try:
					author = author['name']
				except:
					author = author[0]['name']

			form = book_search()

			# Display current books
			if len(book_info) == 0:
				uid = idfrom_username(username)
				with connection.cursor() as cursor:
					global book_info

					cursor.execute("SELECT userid, bookid, name, author, img from wishlist where userid = '%s'" % uid)
					book_info = cursor.fetchall()
				
			notification = notify(idfrom_username(username))

			context = { 'form':form, 'title':result, 'author':author, 'image':img, 'books':book_info, 'flag':flag, 'noti':notification }

			return render(request, "wishlist.html", context)
	else:
		if request.user.is_authenticated():
			global book_info
			form = book_search()
			counts = []
			username = request.user.get_username()
			uid = idfrom_username(username)
			with connection.cursor() as cursor:
				cursor.execute("SELECT name, bookid, author, img from wishlist where userid = '%s'" % uid)
				book_info = cursor.fetchall()
				print "run"
				print(book_info)
				for i in book_info:
					cursor.execute("SELECT COUNT(*) from collection where bookid = '%s' AND userid != '%s' AND available = 1" % (i[1],uid))
					count = cursor.fetchall()
					print count
					counts.append(count[0][0])
			result = ""
			x = len(book_info)
			book_info = list(book_info)
			print len(counts)
			for j in xrange(0,x):
				book_info[j] = list(book_info[j])
			for j in xrange(0,x):
				print j
				book_info[j].append(counts[j])

			notification = notify(idfrom_username(username))
			
			context = { 'form':form, 'title':result, 'books':book_info, 'flag':flag, 'noti':notification }

			return render(request, "wishlist.html", context)
		else:
			return HttpResponseRedirect('/')

def remove_book(request, blist="collection"):
	if request.user.is_authenticated():
		if request.method == "POST":
			bookid = request.POST.get('bookid')
			username = request.user.get_username()
			uid = idfrom_username(username)
			with connection.cursor() as cursor:
				try:
					cursor.execute("DELETE from %s where bookid = '%s' AND userid = '%s'" % (blist,bookid,uid))
				except:
					return HttpResponseRedirect('/')			

				if blist == "collection":
					cursor.execute("DELETE from requests where  bookreq = '%s' AND toid = '%s'" % (bookid, uid))
		else:
			return HttpResponseRedirect('/login/')
	else:
		return HttpResponseRedirect('/')

def owners(request):
	if request.user.is_authenticated():
		username = request.user.get_username()
		uid = idfrom_username(username)
		names = []
		if request.method == "POST":
			bookid = request.POST.get('bookid')

			with connection.cursor() as cursor:
				cursor.execute("SELECT userid,book_cond from collection where bookid = '%s' AND userid != '%s' AND available = 1" % (bookid,uid))
				bookdata = cursor.fetchall()

				cursor.execute("SELECT name,bookid,userid from collection where userid = '%s' AND available = 1" % uid)
				collect = cursor.fetchall()

				for i in bookdata:
					cursor.execute("SELECT name from user_detail where userid = '%s'" % i[0])
					name = cursor.fetchall()
					names.append(name[0][0])
			
			x = len(bookdata)
			bookdata = list(bookdata)
			print names
			for j in xrange(0,x):
				bookdata[j] = list(bookdata[j])
			for j in xrange(0,x):
				bookdata[j].append(names[j])	

			notification = notify(idfrom_username(username))
			
			context = { 'owners':bookdata, 'collect':collect, 'reqbook':bookid, 'noti':notification }

			return render(request, "owners.html", context)
		else:
			return HttpResponseRedirect('/requests/')
	else:
		return HttpResponseRedirect('/')

def requests(request):
	if request.user.is_authenticated():
		username = request.user.get_username()
		uid = idfrom_username(username)
		tnames = []
		fnames = []
		cash = 0
		cont_details = []

		if request.method == "POST":
			bookid = request.POST.getlist('collect')
			reqbook = request.POST.get('req_book')
			cash = request.POST.get('cash')
			toid = request.POST.get('userid')
			books = '_'.join(bookid)

			try:
				cash = int(cash)
			except:
				cash = 0

			print toid
			print "delete"
			with connection.cursor() as cursor:
				# cursor.execute("SELECT userid,book_cond from collection where bookid = '%s'" % bookid)
				# bookdata = cursor.fetchall()
				status = 0
				cursor.execute("""INSERT INTO requests VALUES ('%s','%s','%s','%s','%s','%s','%s')""" % (toid,uid,reqbook,books,cash,status,0))
				
				cursor.execute("""SELECT name, email, contact FROM user_detail where userid = '%s' """ % (toid))
				details = cursor.fetchall()
				name = details[0][0]
				email = details[0][1]
				cont = details[0][2]

				name = name.split(' ')
				name = name[0]

				cursor.execute("""SELECT name FROM collection where bookid = '%s' """ % (reqbook))
				bk_name = cursor.fetchall()
				bk_name = bk_name[0][0]

				message = "Hello, " + name + "!"+ "\n \n" + "You have received a request for your book " + "\"" + bk_name + "\"" + ". Login soon and view the offer.\n" + "\n" + "Cheers!\n" + "Team BookFish"

			print email
			print message


			send_mail('New Book Request', message, 'no-reply@bookfish.in', [email], fail_silently=True,)

			try:
				sendSMS('8nR3v2N0Mn8-kLkdIOfwurKVo16acNGOHC3v9stoF0', cont, 'TXTLCL', message)
			except:
				print "sms not sent"

		with connection.cursor() as cursor:
			# cursor.execute("SELECT userid,book_cond from collection where bookid = '%s'" % bookid)
			# bookdata = cursor.fetchall()

			cursor.execute("SELECT fromid,bookreq,status,books,reqid,money from requests where toid = '%s'" % uid)
			fromid = cursor.fetchall()

			cursor.execute("SELECT toid,bookreq,status,reqid,books,money from requests where fromid = '%s'" % uid)
			toid = cursor.fetchall()

			for i in toid:
				cursor.execute("SELECT userid,contact,email from user_detail where userid = '%s'" % i[0])
				details = cursor.fetchall()
				if details not in cont_details:
					cont_details.append(details)
				cursor.execute("SELECT name from user_detail where userid = '%s'" % i[0])
				name = cursor.fetchall()
				tnames.append(name[0][0])

			x = len(toid)
			y = len(fromid)
			toid = list(toid)
			fromid = list(fromid)

			for j in xrange(0,x):
				toid[j] = list(toid[j])
				if toid[j][2] == 0:
					toid[j][2] = "Awaiting Response"
				elif toid[j][2] == 1:
					toid[j][2] = "Accepted"
				else:
					toid[j][2] = "Rejected"

			for j in xrange(0,y):
				fromid[j] = list(fromid[j])
				if fromid[j][2] == 0:
					fromid[j][2] = "Awaiting Response"
				elif fromid[j][2] == 1:
					fromid[j][2] = "Accepted"
				else:
					fromid[j][2] = "Rejected"

			for i in fromid:
				temp = i[3].split('_')
				if len(i[3]) != 0:
					for j in xrange(0,len(temp)):
						print temp[j]
						with connection.cursor() as cursor:
							cursor.execute("SELECT name from collection where bookid='%s'" % temp[j])
							bk_name = cursor.fetchall()
						print bk_name
						try:
							bk_name = bk_name[0][0]
						except:
							bk_name = "NA"
						temp[j] = bk_name
					i[3] = ", ".join(temp)
				else:
					i[3] = "None"

				with connection.cursor() as cursor:
					cursor.execute("SELECT userid,contact,email from user_detail where userid = '%s'" % i[0])
					details = cursor.fetchall()
					if details not in cont_details:
						cont_details.append(details)
		
					cursor.execute("SELECT name from user_detail where userid = '%s'" % i[0])
					name = cursor.fetchall()

				fnames.append(name[0][0])

			for i in toid:
				temp = i[4].split('_')
				if len(i[4]) != 0:
					for j in xrange(0,len(temp)):
						print temp[j]
						with connection.cursor() as cursor:
							cursor.execute("SELECT name from collection where bookid='%s'" % temp[j])
							bk_name = cursor.fetchall()
						print bk_name
						try:
							bk_name = bk_name[0][0]
						except:
							bk_name = "NA"
						temp[j] = bk_name
					i[4] = ", ".join(temp)
				else:
					i[4] = "None"

		for j in xrange(0,x):
			toid[j].append(tnames[j])

		for j in xrange(0,y):
			fromid[j].append(fnames[j])

		for p in fromid:
			with connection.cursor() as cursor:
				cursor.execute("SELECT name from collection where bookid='%s'" % p[1])
				bk_name = cursor.fetchall()
				bk_name = bk_name[0][0]
			p[1] = bk_name
		
		for p in toid:
			with connection.cursor() as cursor:
				if p[0] == 0:
					cursor.execute("SELECT name from letsbooks where bookid='%s'" % p[1])
				else:
					cursor.execute("SELECT name from collection where bookid='%s'" % p[1])
				bk_name = cursor.fetchall()
				bk_name = bk_name[0][0]
			p[1] = bk_name

		print cont_details
		print toid

		notification = notify(idfrom_username(username))
			
		context = { 'sent_req':toid, 'rec_req':fromid, 'contacts':cont_details, 'noti':notification }

		return render(request, "requests.html", context)

	else:
		return HttpResponseRedirect('/')

def about(request):
	return render(request, "about.html")

def contact(request):
	return render(request, "contact.html")

@csrf_protect
def addbook(request):

	print("Addbook running")
	title = request.POST.get('title')
	ID = data['GoodreadsResponse']['book']['id']

	#Server Connection to MySQL:
	conn = MySQLdb.connect(host= "localhost",user="root",passwd="root",db="startup")
	x = conn.cursor()

	x.execute("""INSERT INTO user_detail VALUES ('%s','%s','%s','%s','%s','%s','%s')"""%(name,0,contact,address,email,username,password))
	conn.commit()

def admin(request):

	if request.method == 'POST':
		username = request.POST.get('user')
		uid = idfrom_username(username)
		delete_acc(username,uid)
	
	with connection.cursor() as cursor:
		cursor.execute("SELECT * from user_detail")
		det = cursor.fetchall()	

		cursor.execute("SELECT * from auth_user")
		auth = cursor.fetchall()

		cursor.execute("SELECT * from collection")
		coll = cursor.fetchall()

	# subject = "Thanks for your support."

	# message = "Dear Bookworms!" + "\n \n" + "Thank you for being a part of BookFish at it's very start. We have had an overwhelming response from you guys and we appreciate you supporting us and making this journey of connecting readers around the campus worthwhile. Your feedback and suggestions help us to improve the platform and we can't thank you enough for that. Hope you guys have seen the latest updates that allow you to get user ratings of books and other details to help you make a better choice." + "\n \n" + "We are happy to announce a new update that will notify you of any requests you receive on your BookFish account through your email. The users who have already signed up do not need to verify their email addresses. But, in case you need to change it -- you can easily do so from the website." + "\n \n" + "Feel free to browse the collection and keep getting new stuff to read. We hope to make your experience as smooth as possible and are working hard for the same. Thanks again for being a part of BookFish. Happy Reading!" + "\n \n" + "With love," + "\n" + "Team BookFish." + "\n \n" + "Click below to unsubscribe from these notifications (But, do you really want to?):"

	# mass_mail(subject, message)

	return render(request, "admin.html", {'det':det, 'auth':auth, 'coll':coll})
		
def lradmin(request, log=0):

	if log == '1':
		if request.user.is_authenticated():
			logout(request)
			return HttpResponseRedirect('/lrmanage/')

	if request.method == 'POST':
		form = login_details(request.POST)

		if form.is_valid():
			username = form.cleaned_data['Username']
			password = form.cleaned_data['Password']

			user = authenticate(username=username, password=password)

			print "none"
			if user != None:
				if username != "letsread":
					return HttpResponseRedirect('/lrmanage/')
				else:
					login(request, user)
			else:
				return HttpResponseRedirect('/lrmanage/')
	
		else:
			with connection.cursor() as cursor:
				cursor.execute("SELECT bookid from letsbooks")
				bid = cursor.fetchall()

			for i in bid:
				bookid = request.POST.get(str(i[0]))

				with connection.cursor() as cursor:
					cursor.execute("UPDATE letsbooks set avail='%s' where bookid='%s'" % (bookid, i[0]))

		with connection.cursor() as cursor:
			cursor.execute("SELECT * from letsbooks")
			det = cursor.fetchall()

			for i in range(0,len(det)):
				cursor.execute("SELECT fromid from requests where bookreq = '%s' AND toid = 0" % (det[i][1]))
				ids = cursor.fetchall()

				# Check for owners
				det = list(det)
				det[i] = list(det[i])
				if len(ids) == 0:
					det[i].append(0)
				else:
					det[i].append(1)

		return render(request, "LRadmin.html", {'det':det})

	else:
		if request.user.is_authenticated():
			username = request.user.get_username()

			if username == "letsread":
				with connection.cursor() as cursor:
					cursor.execute("SELECT * from letsbooks")
					det = cursor.fetchall()

					for i in range(0,len(det)):
						cursor.execute("SELECT fromid from requests where bookreq = '%s' AND toid = 0" % (det[i][1]))
						ids = cursor.fetchall()

						# Check for owners
						det = list(det)
						det[i] = list(det[i])
						if len(ids) == 0:
							det[i].append(0)
						else:
							det[i].append(1)

				return render(request, "LRadmin.html", {'det':det})

			else:
				form = login_details()
				return render(request, "LRadminlogin.html", {'log_form':form})
		else:
			form = login_details()
			return render(request, "LRadminlogin.html", {'log_form':form})

def lrdisp_req(request):

	if request.user.is_authenticated():
		username = request.user.get_username()

		if username == "letsread":
			cont_details = []
			tnames = []

			with connection.cursor() as cursor:
				cursor.execute("SELECT fromid,bookreq,status,reqid from requests where toid = 0")
				fromid = cursor.fetchall()

			y = len(fromid)
			fromid = list(fromid)

			for j in xrange(0,y):
				fromid[j] = list(fromid[j])
				if fromid[j][2] == 0:
					fromid[j][2] = "Awaiting Response"
				elif fromid[j][2] == 1:
					fromid[j][2] = "Accepted"
				else:
					fromid[j][2] = "Rejected"

			for p in fromid:
				with connection.cursor() as cursor:
					cursor.execute("SELECT name from letsbooks where bookid='%s'" % p[1])
					bk_name = cursor.fetchall()
					bk_name = bk_name[0][0]
				p[1] = bk_name

			for i in fromid:
				with connection.cursor() as cursor:
					cursor.execute("SELECT userid,contact,email from user_detail where userid = '%s'" % i[0])
					details = cursor.fetchall()
					if details not in cont_details:
						cont_details.append(details)
					cursor.execute("SELECT name from user_detail where userid = '%s'" % i[0])
					name = cursor.fetchall()
					i.append(name[0][0])

			context = {'rec_req':fromid, 'contacts':cont_details}

			return render(request, "LRreq.html", context)
		else:
			return HttpResponseRedirect('/')
	else:
		return HttpResponseRedirect('/')

def lruser(request, sort="name"):

	flag = 1;

	if request.user.is_authenticated():

		username = request.user.get_username()
		uid = idfrom_username(username) 

		if request.method == 'POST':
			bid = request.POST.get('own_book')

			with connection.cursor() as cursor:
				cursor.execute("SELECT * from requests where fromid='%s' AND bookreq='%s' AND toid=0" % (uid, bid))
				check = cursor.fetchall()

				if len(check) == 0:
					lrreq(uid,bid)
					return HttpResponseRedirect('/requests/')

		try:
			with connection.cursor() as cursor:
				if sort == "name" or sort == "author":
					cursor.execute("SELECT bookid,name,author,pages,rating,img from letsbooks where avail = 1 ORDER BY %s" % (sort))
				else:
					cursor.execute("SELECT bookid,name,author,pages,rating,img from letsbooks where avail = 1 ORDER BY %s desc" % (sort))

				books = cursor.fetchall()
		except:
			return HttpResponseRedirect('/')	

		with connection.cursor() as cursor:
			cursor.execute("SELECT * from requests where toid = 0 AND fromid = '%s' AND (status = 0 OR status = 1)" % (uid))
			check = cursor.fetchall()

			if len(check) != 0:
				flag = 0;

		context = {'books':books,'fl':flag} 

		return render(request,"LRbooks.html",context)
	else:
		return HttpResponseRedirect('/')

def lrreq(uid = "", bid = ""):
	try:
		with connection.cursor() as cursor:
			cursor.execute("INSERT INTO requests values(0,'%s','%s','',0,0,0)" % (uid, bid))
			cursor.execute("UPDATE letsbooks set avail = 0 where bookid = '%s'" % (bid))
	except:
		return HttpResponseRedirect('/')

def change_pass(request, rawid = ""):

	uid = int(rawid)/253
	
	with connection.cursor() as cursor:
		print uid
		cursor.execute("SELECT username FROM user_detail where userid = '%s'" % uid)
		username = cursor.fetchall()
		username = username[0][0]

	if request.method == 'POST':
		form = pass_input(request.POST)

		if form.is_valid():
			password = form.cleaned_data['Password']

			u = User.objects.get(username__exact=username)
			u.set_password(password)
			u.save()

			return HttpResponseRedirect('/')

	else:
		form = pass_input()

	return render(request, "changepass.html", {'rawid':rawid, 'form':form})

def lraccept(request):

	if request.user.is_authenticated():
		username = request.user.get_username()

		if username == "letsread":
			reqid = request.POST.get('reqid')
			with connection.cursor() as cursor:
				cursor.execute("UPDATE requests set status=1 where reqid = '%s'" % reqid)
				cursor.execute("""SELECT fromid, bookreq FROM requests where reqid = '%s' """ % (reqid))
				temp = cursor.fetchall()
				uid = temp[0][0]
				reqbook = temp[0][1]

				cursor.execute("""SELECT name, email, contact FROM user_detail where userid = '%s' """ % (uid))
				details = cursor.fetchall()
				name = details[0][0]
				email = details[0][1]
				cont = details[0][2]

				name = name.split(' ')
				name = name[0]

				cursor.execute("""SELECT name FROM letsbooks where bookid = '%s' """ % (reqbook))
				bk_name = cursor.fetchall()
				bk_name = bk_name[0][0]

			message = "Hello, " + name + "!"+ "\n \n" + "Your request for book " + "\"" + bk_name + "\"" + " has been accepted" + ". Login soon and contact the owner.\n" + "\n" + "Cheers!\n" + "Team BookFish"

			try:
				send_mail('Book Request accepted', message, 'no-reply@bookfish.in', [email], fail_silently=False,)
			except:
				print "mail not sent"

			try:
				sendSMS('8nR3v2N0Mn8-kLkdIOfwurKVo16acNGOHC3v9stoF0', cont, 'TXTLCL', message)
			except:
				print "sms not sent"

			return HttpResponseRedirect('/lrreq/')
		else:
			return HttpResponseRedirect('/')
	else:
		return HttpResponseRedirect('/')

def lrreject(request):

	if request.user.is_authenticated():
		username = request.user.get_username()

		if username == "letsread":
			reqid = request.POST.get('reqid')
			with connection.cursor() as cursor:
				cursor.execute("UPDATE requests set status=-1 where reqid = '%s'" % reqid)

				cursor.execute("""SELECT fromid, bookreq FROM requests where reqid = '%s' """ % (reqid))
				temp = cursor.fetchall()
				uid = temp[0][0]
				reqbook = temp[0][1]

				cursor.execute("UPDATE letsbooks set avail=1 where bookid = '%s'" % reqbook)

				return HttpResponseRedirect('/lrreq/')
		else:
				return HttpResponseRedirect('/')
	else:
		return HttpResponseRedirect('/')

